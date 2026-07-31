import contextlib
import datetime

try:
    from ._imports import without_local_module_shadowing
except ImportError:
    from _imports import without_local_module_shadowing

with without_local_module_shadowing(__file__):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

try:
    from ._layout import render_grid
    from .source import as_source
except ImportError:
    from _layout import render_grid
    from source import as_source


class XLSXExtractor(object):
    def __init__(self, filename, include_hidden=False, formulas=False, name=None):
        self.source = as_source(filename, name)
        self.filename = self.source.name
        self.include_hidden = include_hidden
        self.formulas = formulas
        self.sheets = self._extract()
        self.text = self._render()
        self.provenance = self.source.provenance("xlsx")

    def get_text(self):
        return self.text

    def get_structure(self):
        return {"type": "xlsx", "provenance": self.provenance, "sheets": self.sheets}

    def get_sheets(self):
        return self.sheets

    def _extract(self):
        with contextlib.ExitStack() as stack:
            values_stream = stack.enter_context(self.source.open())
            values_book = load_workbook(
                values_stream, data_only=not self.formulas, read_only=False
            )
            stack.callback(values_book.close)
            if self.formulas:
                formula_book = values_book
            else:
                formula_stream = stack.enter_context(self.source.open())
                formula_book = load_workbook(
                    formula_stream, data_only=False, read_only=False
                )
                stack.callback(formula_book.close)
            sheets = []
            for worksheet in values_book.worksheets:
                if worksheet.sheet_state != "visible" and not self.include_hidden:
                    continue
                formula_sheet = formula_book[worksheet.title]
                sheets.append(self._extract_sheet(worksheet, formula_sheet))
            return sheets

    def _extract_sheet(self, worksheet, formula_sheet):
        populated = [
            cell
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
            or formula_sheet[cell.coordinate].data_type == "f"
        ]
        if not populated:
            return {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "range": None,
                "rows": [],
                "mergedRanges": [],
                "columnWidths": [],
            }

        min_row = min(cell.row for cell in populated)
        max_row = max(cell.row for cell in populated)
        min_col = min(cell.column for cell in populated)
        max_col = max(cell.column for cell in populated)

        merged = []
        merged_anchors = {}
        covered = set()
        for cell_range in worksheet.merged_cells.ranges:
            if (
                cell_range.max_row < min_row
                or cell_range.min_row > max_row
                or cell_range.max_col < min_col
                or cell_range.min_col > max_col
            ):
                continue
            item = {
                "range": str(cell_range),
                "minRow": cell_range.min_row,
                "minCol": cell_range.min_col,
                "rowSpan": cell_range.max_row - cell_range.min_row + 1,
                "colSpan": cell_range.max_col - cell_range.min_col + 1,
            }
            merged.append(item)
            merged_anchors[(cell_range.min_row, cell_range.min_col)] = item
            for row in range(cell_range.min_row, cell_range.max_row + 1):
                for col in range(cell_range.min_col, cell_range.max_col + 1):
                    if (row, col) != (cell_range.min_row, cell_range.min_col):
                        covered.add((row, col))

        rows = []
        for row_index in range(min_row, max_row + 1):
            if worksheet.row_dimensions[row_index].hidden and not self.include_hidden:
                continue
            row_data = []
            for col_index in range(min_col, max_col + 1):
                if worksheet.column_dimensions[
                    get_column_letter(col_index)
                ].hidden and not self.include_hidden:
                    continue
                coordinate = worksheet.cell(row_index, col_index).coordinate
                value_cell = worksheet[coordinate]
                formula_cell = formula_sheet[coordinate]
                value = value_cell.value
                formula = formula_cell.value if formula_cell.data_type == "f" else None
                if value is None and formula:
                    value = formula
                item = {
                    "coordinate": coordinate,
                    "value": self._format_value(value),
                    "formula": formula,
                    "covered": (row_index, col_index) in covered,
                }
                merge = merged_anchors.get((row_index, col_index))
                if merge:
                    item.update(
                        rowSpan=merge["rowSpan"], colSpan=merge["colSpan"]
                    )
                row_data.append(item)
            rows.append(row_data)

        widths = []
        for col_index in range(min_col, max_col + 1):
            letter = get_column_letter(col_index)
            dimension = worksheet.column_dimensions[letter]
            if dimension.hidden and not self.include_hidden:
                continue
            widths.append(float(dimension.width or 13.0))

        return {
            "name": worksheet.title,
            "state": worksheet.sheet_state,
            "range": "{}:{}".format(
                worksheet.cell(min_row, min_col).coordinate,
                worksheet.cell(max_row, max_col).coordinate,
            ),
            "rows": rows,
            "mergedRanges": merged,
            "columnWidths": widths,
        }

    @staticmethod
    def _format_value(value):
        if value is None:
            return ""
        if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
            return value.isoformat()
        return str(value)

    def _render(self):
        rendered = []
        for sheet in self.sheets:
            if not sheet["rows"]:
                rendered.append("[시트: {}] (비어 있음)".format(sheet["name"]))
                continue
            grid = []
            for row in sheet["rows"]:
                values = []
                for cell in row:
                    if cell["covered"]:
                        values.append("")
                        continue
                    value = cell["value"]
                    if cell.get("rowSpan", 1) > 1 or cell.get("colSpan", 1) > 1:
                        value += " (병합 {}×{})".format(
                            cell.get("rowSpan", 1), cell.get("colSpan", 1)
                        )
                    values.append(value)
                grid.append(values)
            rendered.append(
                render_grid(
                    grid,
                    title="[시트: {} {}]".format(sheet["name"], sheet["range"]),
                    column_weights=sheet["columnWidths"],
                )
            )
        return "\n\n".join(rendered)


def get_text(filename):
    print(XLSXExtractor(filename).get_text())


if __name__ == "__main__":
    import sys

    get_text(sys.argv[1])
